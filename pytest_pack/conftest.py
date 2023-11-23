import pytest
import pandas as pd
from openpyxl import load_workbook

@pytest.fixture(scope="class", autouse=True)
def class_set_up():
    print("\n I run before once from conftest")
    yield
    print("I run after once from conftest")

@pytest.fixture
def set_up():
    print("I run before every test")
    yield
    print("I run after every test")

# @pytest.mark.parametrize()

@pytest.fixture
def data_load():
    return ["Pradeep", "Ramola", "pr@gmai.com"]

@pytest.fixture(params=[("chrome", "pradeep"), ("ff", "ramola"), ("IE", "Anku")])
def cross_browser(request):
    return request.param


@pytest.fixture
def login(request):
    return request.param

def get_user_credentials():
    excel_data = pd.read_excel("login.xlsx")
    return excel_data.to_dict(orient='records')

@pytest.fixture(params=get_user_credentials())
def user_credentials(request):
    return request.param

# @pytest.fixture
# def excel_data():
#     # Read the Excel file with user credentials
#     return pd.read_excel('login.xlsx')
#
# @pytest.fixture(params=range(3))  # Assuming 3 sets of credentials in the Excel file
# def user_credentials(excel_data, request):
#     return excel_data.iloc[request.param]



def get_user_creds_xl():
    wb = load_workbook("login.xlsx")
    sheet = wb.active
    creds = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        user, password = row
        creds.append({"Username": user, "Password": password})

    return creds


@pytest.fixture(params=get_user_creds_xl())
def user_creds_xl(request):
    return request.param



